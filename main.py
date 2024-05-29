# This cmd program scraps a Youtube playlist
# and writes the name, artist/uploder and link
# into an new or existing Excel file


# necessary libraries: selenium, openpyxl
from sys import exit
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def check_input(url):
    match url:
        case s if s.startswith("https://www.youtube.com/playlist?list="):
            print("Valid Youtube link, continuing...")
            print()
            return 0
        case s if s.startswith("https://music.youtube.com/playlist?list="):
            print("Valid Youtube-Music link, continuing...")
            print()
            return 1
        case "-q":
            print("Quitting the program.\nGoodbye...")
            exit(1)
        case "-h":
            print("""
            This cmd program scraps a Youtube playlist
            and writes the name, artist/uploder and link
            into an new or existing Excel file.
    
            Only works with links like this: "https://www.youtube.com/playlist?list=".
            Or YT-music playlists like this: "https://music.youtube.com/playlist?list=".
             
             My commands like -h for help and -q to quit are case sensitive.
             I don't feel like implementing that rn ngl.
             
             I'm just having fun with this and tr
             Thank you for using my program :3\n
            """)
            return 8
        case _:
            print("\nInvalid input. Please try again or press -h for help.\n")
            return 9


def get_normal_info(driver):
    # Cookie & load complete page
    driver.find_element(By.XPATH, '/html/body/c-wiz/div/div/div/div[2]/div[1]/div[3]/div[1]/form[2]/div').click()
    # driver.find_element(By.TAG_NAME, "body").send_keys(Keys.CONTROL+Keys.END)
    html = driver.find_element(By.TAG_NAME, 'html')
    while True:
        s1 = driver.page_source
        html.send_keys(Keys.END)
        sleep(1)
        html.send_keys(Keys.END)
        s2 = driver.page_source
        if s1 == s2:
            break

    # The thing that actually does stuff:

    content = driver.find_element(By.ID, "contents")

    results = content.find_elements(By.CLASS_NAME, "style-scope.ytd-playlist-video-list-renderer")

    vids = results[4:(len(results) - 1)]

    print("-" * 20)

    entries = []

    for entry in vids:
        try:
            new_entry = []

            entry_text = entry.text.splitlines()
            new_entry.append(entry_text[2:4])

            try:
                link_pos = entry.find_element(By.CLASS_NAME, "yt-simple-endpoint.style-scope.ytd-playlist-video-renderer")
                link = link_pos.get_attribute('href').split('&')
                new_entry.append(link[0])

            except:
                new_entry.append("unavailable")

            print(new_entry)
            entries.append(new_entry)

        except Exception as error:
            new_entry = [error, "unavailable", "unavailable"]

    return entries

def get_music_info(driver):
    # Cookie & load complete page
    driver.find_element(By.XPATH, '/html/body/c-wiz/div/div/div/div[2]/div[1]/div[3]/div[1]/form[2]/div').click()
    #driver.find_element(By.TAG_NAME, "body").send_keys(Keys.CONTROL+Keys.END)
    html = driver.find_element(By.TAG_NAME, 'html')
    while True:
        s1 = driver.page_source
        sleep(1)
        html.send_keys(Keys.PAGE_DOWN)
        html.send_keys(Keys.END)
        sleep(1)
        html.send_keys(Keys.END)
        s2 = driver.page_source
        if s1 == s2:
            break

    # The thing that actually does stuff:

    content = driver.find_element(By.ID, "contents")

    results = content.find_elements(By.CLASS_NAME, "style-scope.ytmusic-playlist-shelf-renderer")

    vids = results[2:(len(results) - 1)]

    print("-" * 20)

    entries = []

    for entry in vids:
        try:
            new_entry = []

            entry_text = entry.text.splitlines()
            new_entry.append(entry_text[0:2])

            try:
                link_pos = entry.find_element(By.CLASS_NAME, "yt-simple-endpoint.style-scope.yt-formatted-string")
                link = link_pos.get_attribute('href').split('&')
                new_entry.append(link[0])

            except:
                new_entry.append("unavailable")

            print(new_entry)
            entries.append(new_entry)

        except Exception as error:
            new_entry = [error, "unavailable", "unavailable"]

    return entries


def search_for_point(list):
    point = "•"

    for i in range(len(list)):
        if list[i] == point:
            return [(i - 2), (i - 1)]


def get_information(driver, mode):

    pathToResults = ["style-scope.ytd-playlist-video-list-renderer", "style-scope.ytmusic-playlist-shelf-renderer"]
    startOfActualResults = [4, 2]
    actualLinkPos = ["yt-simple-endpoint.style-scope.ytd-playlist-video-renderer", "yt-simple-endpoint.style-scope.yt-formatted-string"]


    # Cookie & load complete page
    driver.find_element(By.XPATH, '/html/body/c-wiz/div/div/div/div[2]/div[1]/div[3]/div[1]/form[2]/div').click()
    # driver.find_element(By.TAG_NAME, "body").send_keys(Keys.CONTROL+Keys.END)
    html = driver.find_element(By.TAG_NAME, 'html')
    while True:
        s1 = driver.page_source
        sleep(1)
        html.send_keys(Keys.PAGE_DOWN)
        html.send_keys(Keys.END)
        sleep(1)
        html.send_keys(Keys.END)
        s2 = driver.page_source
        if s1 == s2:
            break

    # The thing that actually does stuff:

    content = driver.find_element(By.ID, "contents")

    results = content.find_elements(By.CLASS_NAME, pathToResults[mode])

    vids = results[startOfActualResults[mode]:(len(results) - 1)]

    print("-" * 20)

    entries = []

    for entry in vids:
        try:
            new_entry = []

            entry_text = entry.text.splitlines()

            match mode:
                case 0:
                    actualEntryText = search_for_point(entry_text)
                    new_entry.append(entry_text[actualEntryText[0]])
                    new_entry.append(entry_text[actualEntryText[1]])
                case 1:
                    new_entry.append(entry_text[0])
                    new_entry.append(entry_text[1])

            try:
                link_pos = entry.find_element(By.CLASS_NAME, actualLinkPos[mode])
                link = link_pos.get_attribute('href').split('&')
                new_entry.append(link[0])

            except:
                new_entry.append("unavailable")

            print(new_entry)
            entries.append(new_entry)

        except Exception as error:
            entries = [error, "unavailable", "unavailable"]

    return entries


def create_excel_with_table(data, filename):
    # Ensure the filename has the correct extension
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"

    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Define the headers
    headers = ["Name", "Uploader", "Link"]

    # Define styles
    header_font = Font(size=14, bold=True)

    # Write the headers to the first row with styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font

    # Write the data to the subsequent rows without alternating row colors
    for row_num, row_data in enumerate(data, start=2):
        for col_num, cell_value in enumerate(row_data, start=1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    # Set filters on the headers
    ws.auto_filter.ref = ws.dimensions

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook to a file
    wb.save(filename)
    print(f"Data written to {filename}")


def main():
    print("Welcome to YT-Playlist2Excel")
    filename = input("Please enter a filename: ")

    while True:
        #url = input("Please input your playlist link, press -h for help or -q to quit: ")
        # Test URL
        url = "https://music.youtube.com/playlist?list=PLdxfyzVbTnHn_7pgh4twqXGQva2EgfnF7"
        #url = "https://www.youtube.com/playlist?list=PLdxfyzVbTnHl3sJ5mhYTULvKlj4q0HIDY"

        mode = check_input(url)
        if mode < 5:
            break

    # Init Selenium
    driver = webdriver.Firefox()
    driver.get(url)
    driver.implicitly_wait(10)

    # Getting the information

    entries = get_information(driver, mode)

    print(entries)
    print(str(len(entries)) + " Videos found.")
    driver.close()

    # Save in Excel
    create_excel_with_table(entries, filename)

    print("Finished...")


if __name__ == "__main__":
    main()
